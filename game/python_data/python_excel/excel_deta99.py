from openpyxl import Workbook, load_workbook


class PracticeExcel:

    def create(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "身高"
        ws['B1'] = "体重"
        # # 身高
        self.height = [180, 160, 170, 155]
        # # 体重
        weight = [66, 50, 40, 30]
        for i in range(len(self.height)):
            ws.cell(row=2 + i, column=1, value=self.height[i])
            ws.cell(row=2 + i, column=2, value=weight[i])

        # Save the file
        wb.save("sample.xlsx")

    def health_weight(self):
        ld = load_workbook(filename="sample.xlsx")
        sheet = ld.active
        sheet["C1"] = "备注"
        for i in range(len(self.height)):
            height = sheet.cell(row=2 + i, column=1).value
            weight = sheet.cell(row=2 + i, column=2).value
            # 获取身高对应的健康体重
            health_w = (height-70)*0.6
            if weight == health_w:
                print("这是健康的体重", weight)
                sheet.cell(row=2 + i, column=3).value = "健康体重"

        ## 注意，一定要保存
        ld.save("sample1.xlsx")



pe = PracticeExcel()
pe.create()
pe.health_weight()