from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = "身高"
ws['B1'] = "体重"
#身高
height=[180,160,170,155]
# 体重
weight=[60,50,40,30]

# Save the file
# 注意变量i是从0开始的
for  i  in range(len(height)):
    # 可以打印一下看一下
    print(i)
    ws.cell(row=2+i,column=1,value=height[i])
    ws.cell(row=2+i,column=2,value=height[i])

wb.save("sample.xlsx")