from openpyxl import Workbook, load_workbook


class PracticeExcel:

    def create(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "身高"
        ws['B1'] = "体重"
        # 身高
        self.height = [180, 160, 170, 155]
        # 体重
        weight = [60, 50, 40, 30]
        # Save the file
        # 注意变量i是从0开始的
        for i in range(len(self.height)):
            # 可以打印一下看一下
            # print(i)
            ws.cell(row=2 + i, column=1, value=self.height[i])
            ws.cell(row=2 + i, column=2, value=weight[i])
        wb.save("sample.xlsx")

    def health_weight(self):
        # filename是文件名，文件名是sample.xlsx
        # 读取数据也是库里的一个类？load_workbook
        ld = load_workbook("filename=sample.xlsx")
        # sheet是看出来的。看创建的第一个页签名字。因是通过名字找对应页签。或sheet=ld.active也可以，也是库里的方法。
        sheet=ld.active
        sheet["c1"]="备注"
        # 先拿到value值，因等下要准备计算。用循环方式拿
        for i in range(len(self.height)):
            height=sheet.cell(row=2 + i, column=1).value
            weight=sheet.cell(row=2 + i, column=2).value
            # 接下去要去做判断,先获取身高对应的健康体重
            health_w=(height-70)*0.6
            if weight==health_w:
                print("这是健康的体重",weight)
                sheet.cell(row=2 + i, column=3).value="健康体重"
#       注意一定切记保存要保存（在整个代码外保存，不是写在循环里(文件名重新写一个)
        ld.save("sample1.xlsx")

pe=PracticeExcel()
pe.create()
pe.health_weight()

